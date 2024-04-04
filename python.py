import urllib.request
from bs4 import BeautifulSoup
import re
import openpyxl
import pyinputplus as pyip
import time

def scrapeProducts(url):
    try:
        # Create a request object with a Chrome User-Agent header
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36'
        }
        req = urllib.request.Request(url, headers=headers)

        # Open the URL and fetch HTML content
        with urllib.request.urlopen(req) as response:
            htmlContent = response.read()

        # Parse the HTML content
        soup = BeautifulSoup(htmlContent, 'html.parser')

        # Find all product names
        productNameElems = soup.find_all('a', class_='product-slab__title')
        if productNameElems:
            productNames = [elem.text.strip() for elem in productNameElems]
        else:
            print("Product names not found.")
            return [], []

        # Find all product prices using regular expression
        productPriceElems = soup.find_all(string=re.compile(r'\$\d+\.\d+'))
        if productPriceElems:
            productPrices = [price.strip() for price in productPriceElems]
        else:
            print("Product prices not found.")
            return [], []

        return productNames, productPrices

    except Exception as e:
        print("An error occurred:", str(e))
        return [], []

def write_to_excel(products, prices):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Report"

    # Write product information to excel
    ws.append(["Product Name", "Price"])
    for product, price in zip(products, prices):
        ws.append([product, price])

    wb.save("report.xlsx")
    print("Report has been successfully generated.")

def menu(products, prices):
    exit_flag = False

    while not exit_flag:
        print("\nMenu:")
        print("1. Update any cell in the spreadsheet")
        print("2. Delete any row")
        print("3. View a range of cells")
        print("4. Exit")
        
        choice = pyip.inputMenu(['1', '2', '3', '4'], numbered=True)
        
        if choice == '1':
            print("Updating any cell in the spreadsheet")
            update_cell(products, prices)
        elif choice == '2':
            print("Deleting any row")
            delete_row(products, prices)
        elif choice == '3':
            print("Viewing a range of cells")
            view_range()
        elif choice == '4':
            print("Exiting...")
            exit_flag = True

def update_cell(products, prices):
    try:
        wb = openpyxl.load_workbook("report.xlsx")
        ws = wb.active

        row = int(input("Enter the row number to update: "))
        col = input("Enter the column letter to update (e.g., A, B, C, ...): ")
        new_value = input("Enter the new value: ")

        # Check if the row and column are valid
        if 1 <= row <= ws.max_row and 1 <= openpyxl.utils.column_index_from_string(col.upper()) <= ws.max_column:
            # Update the cell value
            ws[f"{col.upper()}{row}"] = new_value
            # Save the current time
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            # Write the time to a separate column (e.g., column Z)
            ws[f"D{row}"] = current_time
            wb.save("report.xlsx")
            print("Cell updated successfully.")
        else:
            error_message = "Invalid row or column."
            # Write the error message to a file
            with open("error_log.txt", "w") as f:
                f.write(error_message)
            print(error_message)

    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        # Write the error message to a file
        with open("error_log.txt", "w") as f:
            f.write(error_message)
        print(error_message)


def delete_row(products, prices):
    try:
        wb = openpyxl.load_workbook("report.xlsx")
        ws = wb.active

        row = int(input("Enter the row number to delete: "))

        # Check if the row is valid
        if row <= ws.max_row:
            ws.delete_rows(row)
            wb.save("report.xlsx")
            print("Row deleted successfully.")
        else:
            print("Invalid row number.")

    except Exception as e:
        print("An error occurred:", str(e))

def view_range():
    try:
        wb = openpyxl.load_workbook("report.xlsx")
        ws = wb.active

        start_row = int(input("Enter the starting row number: "))
        end_row = int(input("Enter the ending row number: "))
        start_col = input("Enter the starting column letter (e.g., A, B, C, ...): ")
        end_col = input("Enter the ending column letter (e.g., A, B, C, ...): ")

        for row in range(start_row, end_row + 1):
            for col in range(openpyxl.utils.column_index_from_string(start_col), openpyxl.utils.column_index_from_string(end_col) + 1):
                cell_value = ws.cell(row=row, column=col).value
                print(cell_value, end="\t")
            print()

    except Exception as e:
        print("An error occurred:", str(e))


# Define the URL of the web page you want to scrape
url = 'https://www.optimumnutrition.com/en-us/Products/c/1000'

# Call the scrapeProducts function with the URL
productNames, productPrices = scrapeProducts(url)

# Print the extracted information
for name, price in zip(productNames, productPrices):
    print("Product Name:", name)
    print("Product Price:", price)
    print()  # Adding a blank line for readability

# Write the extracted information to an Excel file
write_to_excel(productNames, productPrices)

# Call the menu function
menu(productNames, productPrices)
